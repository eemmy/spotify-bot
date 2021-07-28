# -*- coding: utf-8 -*-
import openpyxl
from time import sleep
from pathlib import Path
from random import randint
from selenium import webdriver
from os.path import isfile
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.proxy import Proxy, ProxyType

class WebCrawler:
    # Class constructor
    def __init__(self, listen_times):
        # Get the users accounts
        self.get_accounts()

        # Stores the random times to use in human user simulations
        self.random_times = {
            "pages": [3, 12],
            "musics": [8, 14]
        }

        # Stores spotify routes map
        self.map = {
            'home': 'https://www.spotify.com/br/'
        }

        # Times to listen the music
        self.listen_times = 10000000000000000000000000000000000000000000 if listen_times == 0 else listen_times

    # Wait a random time
    def wait(self, type="pages"):
        # Sleeps in a random time, present in the range of passed type
        sleep(randint(self.random_times[type][0], self.random_times[type][1]))

    # Gets the user accounts
    def get_accounts(self, e=''):
        try:
            # get the file path
            xlsx_file = Path('files', 'contas'+e+'.xlsx')

            # Load the openpyxl object
            wb_obj = openpyxl.load_workbook(xlsx_file)

            # Read the active sheet
            sheet = wb_obj.active

            # Return the recovered accounts
            return sheet
        except:
            print('contas'+e+'.xlsx')
            print("exit")
            # Exit app
            exit()

    # Gets the musics list
    def get_musics(self):
        # get the file path
        xlsx_file = Path('files', 'Músicas.xlsx')

        # Load the openpyxl object
        wb_obj = openpyxl.load_workbook(xlsx_file) 

        # Read the active sheet
        sheet = wb_obj.active

        # Return the recovered data
        return sheet

    # Gets the playlists list
    def get_playlists(self):
        # get the file path
        xlsx_file = Path('files', 'playlists.xlsx')

        # Load the openpyxl object
        wb_obj = openpyxl.load_workbook(xlsx_file) 

        # Read the active sheet
        sheet = wb_obj.active

        # Return the recovered data
        return sheet

    # Gets the proxies
    def get_proxy_list(self):
        # Stores the proxies
        proxies = list()

        # Open the proxy.txt file
        with open('files/proxy.txt') as f:
            # Get the lines from the file
            lines = f.readlines()

            # Loop file lines
            for line in lines:
                # Append the proxy address to proxies
                proxies.append(line.strip())

        # Return the proxies
        return proxies

    # Retuns the proxys list
    def get_proxy(self):
        # Get the configured proxies
        proxies = self.get_proxy_list()

        # Stores the used random proxy
        index = randint(0, len(proxies)-1)

        # Return a random proxy in the list
        return proxies[index]

    # Starts a chromedriver
    def start_webdriver(self):
        # Starts a chrome_options object
        chrome_options = webdriver.ChromeOptions()

        # Change user-agent
        chrome_options.add_argument('--user-agent="Mozilla/5.0 (X11; Linux x86_64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/91.0.4472.106 Safari/537.36"')

        # Verifies if files/proxy.txt
        if (isfile("files/proxy.txt")):
            # Get a proxy from sheet
            proxy = self.get_proxy()

            # Adds the proxy to chrome_options
            chrome_options.add_argument('--proxy-server=%s' % proxy)

            # Show to the user which proxy the bot is using now
            print("\nUsando o proxy: " + proxy)

        # Instantiathe the chromedriver
        self.browser = webdriver.Chrome(chrome_options=chrome_options, executable_path=Path("driver/" 'chromedriver.exe'))

        # Maximize the window
        self.browser.maximize_window()

    # Click in a provided element
    def click(self, el):
        # Click in element using JavaScript Engine
        self.browser.execute_script("arguments[0].click();", el)

    # Authenticates on spotify
    def login(self, email, password):
        # Gets the spotify home page
        self.browser.get(self.map['home'])

        # Wait a random time in page
        self.wait()

        # Get the login button, and click him
        self.click(self.browser.find_element_by_xpath("//a[@data-ga-action='log-in']"))

        # Waits a random time to load the page, and a time inside the page
        sleep(35)
        self.wait()

        # Send the user data to the email and password fields
        self.browser.find_element_by_xpath("//input[@id='login-username']").send_keys(email)
        self.browser.find_element_by_xpath("//input[@id='login-password']").send_keys(password)

        # Click in login button
        self.click(self.browser.find_element_by_xpath("//button[@id='login-button']"))

        # Waits the page load
        sleep(35)

    # Search by a music
    def search_music(self, name, author):
        # Click in search button
        self.click(self.browser.find_element_by_xpath("//a[@href='/search']"))

        # Waiths a random time
        self.wait()

        # Mount the search string
        search_string = author + " " + name

        # Clear search field and send search_string variable to the search field
        self.browser.find_element_by_xpath("//input[@data-testid='search-input']").clear()
        self.browser.find_element_by_xpath("//input[@data-testid='search-input']").send_keys(search_string)

        # Waits a random time
        self.wait()
        sleep(35)

        # Verifies if was multiple play buttons
        if type(self.browser.find_element_by_xpath("//button[@data-testid='play-button']")) is webdriver.remote.webelement.WebElement:
            # Click in the play button
            self.click(self.browser.find_element_by_xpath("//button[@data-testid='play-button']"))
        else:
            # Click in the play button
            self.click(self.browser.find_element_by_xpath("//button[@data-testid='play-button']")[0])

        # Waits the music start
        sleep(20)


    # Pause the script executation during music time
    def listen_music(self):
        # Waits music start
        sleep(6)

        # Infinite loop
        while True:
            # Gets the text from the progress bar
            current_time = self.browser.find_element_by_xpath("//div[contains(@class, 'playback-bar__progress-time')]").text

            # Verifies if the music reached end
            if current_time == "0:00":
                # Exit the loop
                break
            else:
                # Sleep 1 second
                sleep(1)

        # Sleeps a random time before back to music
        self.wait("musics")
        sleep(6)

    # Automate the target, using musics
    def run_music(self, email, password):
        # Get the musics list
        musics = self.get_musics()

        # Starts the webdriver
        self.start_webdriver()

        # Authenticate on spotify
        self.login(email, password)

        # Musics loop controler
        e = 1

        # Loop musics
        while True:
            # Get the current music info
            name = musics["A" + str(e)].value
            author = musics["B" + str(e)].value

            # Verifies if the music name is empty
            if name == None:
                # Exit musics loop
                break

            # Search by the current music
            self.search_music(name, author)

            # Waits the music ends
            self.listen_music()

            # Loop to listen in repeat
            for x in range(1, self.listen_times - 1):
                # Verifies if was multiple play buttons
                if type(self.browser.find_element_by_xpath("//button[@data-testid='play-button']")) is webdriver.remote.webelement.WebElement:
                    # Click in the play button
                    self.click(self.browser.find_element_by_xpath("//button[@data-testid='play-button']"))
                else:
                    # Click in the play button
                    self.click(self.browser.find_element_by_xpath("//button[@data-testid='play-button']")[0])

                # Listen the track
                self.listen_music()

            # Increment musics loop controller
            e += 1

        # Closes the browser
        self.browser.close()

    # Automate the target, using musics inside playlists
    def run_playlist(self, email, password):
        # Get the musics list
        playlists = self.get_playlists()

        # Starts the webdriver
        self.start_webdriver()

        # Authenticate on spotify
        self.login(email, password)

        # Loop controller
        e = 1

        # Loop for playlists
        while True:
            # Get the playlist link from sheet
            playlist_link    = playlists["A" + str(e)].value
            musics_to_listen = playlists["B" + str(e)].value

            # Verifies if the playlist link is empty
            if playlist_link == None:
                # Exit the loop
                break

            # Splits musics_to_listen in a list
            musics_to_listen = musics_to_listen.split(",")

            # Get the playlist page in browser
            self.browser.get(playlist_link)

            # Pause to wait page loads
            sleep(35)

            # Loop to check if the playlist was liked
            for i in range(0, 1):
                try:
                    try:
                        # Get the liked 
                        self.browser.find_element_by_xpath("//button[contains(@class, 'a66d8d62fe56eed3e660b937a9be8a93-scss')]")
                    except:
                        pass
                # Case selenium cant find element, it will throw an exception, handled here
                except:
                    # Click the first fav buttt
                    self.click(self.browser.find_element_by_xpath("//button[@class='_07bed3a434fa59aa1852a431bf2e19cb-scss']"))
                    break

            # Waits a random time
            self.wait()

            # Get the musics list
            musics = self.browser.find_elements_by_xpath("//div[@class='_6b1ff8eab07810e2b7845ffe28430e38-scss']")

            # Loop in musics
            for music_index, music in enumerate(musics):
                # Loop to get the music position
                for x in range(0, 1):
                    # Get the music position
                    try:
                        music_position = musics[music_index].find_element_by_xpath("./child::span").text
                    except: 
                        music_position = str(music_index + 1)

                # Get the play button
                play_button = musics[music_index].find_element_by_xpath("./child::button")

                # Verifies if the current music was defined to be listened
                if music_position in musics_to_listen:
                    # Click the play button
                    self.click(play_button)

                    # Waits the music start
                    sleep(6)

                    # Listen the music
                    self.listen_music()

                    # Loop to listen in repeat
                    for x in range(0, self.listen_times - 1):
                        # Waits a random time
                        self.wait("musics")

                        # Click in the previous track play_button
                        self.click(play_button)

                        # Waits the music start
                        sleep(6)

                        # Listen the music till the end
                        self.listen_music()
                else:
                    # Jump to the next loop executation
                    continue

            # Increments loop controller
            e += 1

        # Closes the browser
        self.browser.close()
